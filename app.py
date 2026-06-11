"""
Descargador automático de cartas glosa — Seguros Mundial
Flujo: login + agregar cartas + descarga de ZIPs del portal.
Estructura final: un solo ZIP por IPS con carpetas DEV/ y LIQ/ + Excel.
"""

import os
import re
import json
import csv
import time
import threading
import logging
import zipfile
import tempfile
import shutil
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_from_directory
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("openpyxl no instalado. No se generara el archivo Excel.")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

BASE_DIR = Path(__file__).parent
port = int(os.environ.get("PORT", 8080))

# En Railway se monta un volumen persistente en /data.
# Localmente (sin volumen) usa la carpeta downloads/ del proyecto.
_data_root = Path(os.environ.get("DATA_DIR", "/data"))
if _data_root.exists():
    DOWNLOAD_DIR = _data_root / "downloads"
else:
    DOWNLOAD_DIR = BASE_DIR / "downloads"
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

LOGIN_URL = "https://a2m-mundial.iqdigital.com.co/ATF_Site/"
CONSULTA_CARTAS_URL = "https://a2m-mundial.iqdigital.com.co/ATF_Site/wallet/settlement-letters"

# ==================== MAPA DE IPS POR NIT ====================
MAPA_IPS = {
    "900267064": "INVERSIONES_AZALUD_CLINICA_BAHIA",
    "900827065": "CENTRO_DE_DIAGNOSTICO_E_IMAGENES_BAHIA",
    "900657731": "CENTRO_MEDICO_Y_DE_REHABILITACION_BAHIA",
    "900826509": "RED_DE_URGENCIAS_DEL_MAGDALENA",
    "900513306": "FUNDACION_MARIA_REINA",
    "900600550": "INVERSIONES_MEDICAS_BARU",
    "900954800": "CENTRO_MEDICO_Y_DE_REHABILITACION_BARU",
    "900631361": "INVERSIONES_MEDICAS_VALLESALUD",
    "900257333": "ODONTOTRANS",
    "901081281": "URGETRAUMA",
    "900792417": "RED_DE_URGENCIAS_DE_LA_COSTA_PACIFICA",
    "901959993": "CLINICA_CORDIALIDAD",
    "900002780": "FUNDACION_CAMPBELL",
    "901523868": "MOVID_IPS_SAS",
    "901057487": "TECNOLOGIA_DIAGNOSTICA_DEL_VALLE",
    "900558595": "FUNDACION_MEDICA_CAMPBELL",
    "901149757": "UNIDAD_MEDICA_DE_TRAUMA_VALLE_SALUD",
    "900900754": "CLINICA_VALLE_SALUD_SAN_FERNANDO",
    "900469882": "CENTRO_MEDICO_SERVISALUD_INTEGRAL_IPS_SAS",
    "802024329": "RED_DE_URGENCIA_DE_LA_COSTA_LTDA",
    "900847382": "CENTRO_MEDICO_Y_DE_REHABILITACION_VALLE_SALUD",
}

def resolver_ips_por_usuario(usuario: str):
    m = re.search(r"(\d{9,12})", usuario or "")
    nit = m.group(1) if m else None
    nombre = MAPA_IPS.get(nit, "IPS_DESCONOCIDA") if nit else "IPS_DESCONOCIDA"
    return nit, nombre

def detectar_tipo_solicitud(valor: str):
    v = (valor or "").strip().upper()
    if v.startswith("DEV"):
        return "No Dev/Obj"
    if v.startswith("LIQ"):
        return "No Liquidación"
    if v.startswith("CMVIQ") or v.startswith("CMV"):
        return "No Radicado"
    return "No Radicado"

# ==================== ESTADO GLOBAL ====================
job_state = {
    "running": False,
    "stopping": False,
    "logs": [],
    "stats": {"total": 0, "descargadas": 0, "errores": 0},
    "finished": False,
    "error": None,
    "errores_detalle": [],
    "descargas_exitosas": [],
}
job_lock = threading.Lock()
current_browser = None
current_context = None
current_dl_dir = None
current_lote = None
current_ips_nombre = None

# ==================== LOGGING CON COLORES ====================
def log(msg, level="info"):
    ts = datetime.now().strftime("%H:%M:%S")
    colors = {
        "info": "\033[94m",
        "success": "\033[92m",
        "warn": "\033[93m",
        "error": "\033[91m",
    }
    reset = "\033[0m"
    color = colors.get(level, colors["info"])
    if level == "success":
        print(f"{color}[{ts}] {msg}{reset}")
    elif level == "error":
        print(f"{color}[{ts}] ERROR: {msg}{reset}")
    elif level == "warn":
        print(f"{color}[{ts}] ADVERTENCIA: {msg}{reset}")
    else:
        print(f"{color}[{ts}] {msg}{reset}")
    
    entry = {"ts": ts, "msg": msg, "level": level}
    with job_lock:
        job_state["logs"].append(entry)
    if level == "error":
        logger.error(msg)
    else:
        logger.info(msg)

def reset_state():
    with job_lock:
        job_state["running"] = False
        job_state["stopping"] = False
        job_state["logs"] = []
        job_state["stats"] = {"total": 0, "descargadas": 0, "errores": 0}
        job_state["finished"] = False
        job_state["error"] = None
        job_state["errores_detalle"] = []
        job_state["descargas_exitosas"] = []

def stop_job():
    global current_browser
    with job_lock:
        job_state["stopping"] = True
    log("Solicitando detencion del proceso...", "warn")
    if current_browser:
        try:
            current_browser.close()
            log("  -> Navegador cerrado por solicitud de stop.")
        except Exception as e:
            log(f"  -> Error al cerrar navegador: {e}", "error")
    generar_zip_parcial()  # ya no se usa en la nueva estructura, pero se mantiene por compatibilidad

def generar_zip_parcial():
    # Esta función se puede eliminar o mantener como respaldo
    pass

# ==================== PERSISTENCIA ====================
def cargar_progreso(ips_dir):
    p = ips_dir / "progreso.json"
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
            completadas = data.get("completadas", [])
            return set(completadas) if isinstance(completadas, list) else set()
        except Exception as e:
            log(f"Error al leer progreso: {e}", "warn")
    return set()

def guardar_progreso(ips_dir, completadas):
    p = ips_dir / "progreso.json"
    try:
        with open(p, "w", encoding="utf-8") as f:
            json.dump({"completadas": list(completadas), "actualizado": datetime.now().isoformat()}, f, indent=2)
    except Exception as e:
        log(f"Error al guardar progreso: {e}", "warn")

# ==================== UTILIDADES PLAYWRIGHT ====================
def _texto_pagina(page):
    try:
        return (page.evaluate("() => document.body ? document.body.innerText : ''") or "")
    except Exception:
        return ""

def _esperar_texto(page, regex, timeout=30, intervalo=0.5):
    fin = time.time() + timeout
    pat = re.compile(regex, re.I)
    while time.time() < fin:
        if job_state.get("stopping"):
            return False
        if pat.search(_texto_pagina(page)):
            return True
        time.sleep(intervalo)
    return False

def _click_por_texto(page, texto, exacto=False, timeout=8000):
    estrategias = []
    if exacto:
        estrategias.append(lambda: page.get_by_text(texto, exact=True).first)
    estrategias.append(lambda: page.get_by_text(texto).first)
    estrategias.append(lambda: page.locator(f"text={texto}").first)
    estrategias.append(lambda: page.get_by_role("button", name=re.compile(re.escape(texto), re.I)).first)
    for estrategia in estrategias:
        try:
            loc = estrategia()
            loc.click(timeout=timeout)
            return True
        except Exception:
            continue
    try:
        clicked = page.evaluate(
            """([t, ex]) => {
                const norm = s => (s||'').toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g,'').trim();
                const target = norm(t);
                const els = document.querySelectorAll('button, a, span, div, li, td, [role="button"]');
                for (const el of els) {
                    const txt = norm(el.textContent);
                    if (ex ? (txt === target) : txt.includes(target)) {
                        if (txt.length <= target.length + 40) { el.click(); return true; }
                    }
                }
                return false;
            }""",
            [texto, exacto],
        )
        return bool(clicked)
    except Exception:
        return False

# ==================== LOGIN ====================
def _hacer_login(page, usuario, password, login_timeout):
    log("Abriendo portal de Seguros Mundial...")
    page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=60000)
    time.sleep(2)

    try:
        select_locator = page.locator('mat-select[name="typeAccess"]').first
        select_locator.click(timeout=5000)
        log("  Clic en desplegable 'Tipo de Acceso'.")
        time.sleep(0.5)
        page.keyboard.press("ArrowDown")
        time.sleep(0.3)
        page.keyboard.press("Enter")
        log("  Teclas: Flecha Abajo + Enter para seleccionar 'Usuario'.")
        time.sleep(0.5)
    except Exception as e:
        log(f"  No se pudo seleccionar 'Usuario' con teclado: {e}.", "warn")
        try:
            page.locator('mat-option, .mat-option', has_text=re.compile(r"^Usuario$", re.I)).first.click(timeout=5000)
            log("  Opción 'Usuario' seleccionada por clic directo (fallback).")
        except Exception as e2:
            log(f"  Fallback también falló: {e2}", "warn")

    try:
        page.fill('input[name="user"]', usuario, timeout=5000)
        log("  Campo 'Usuario' llenado.")
    except Exception:
        page.locator('input:not([type="password"])').first.fill(usuario)
        log("  Campo 'Usuario' llenado (fallback).")

    try:
        page.fill('input[name="password"]', password, timeout=5000)
        log("  Campo 'Contraseña' llenado.")
    except Exception:
        page.locator('input[type="password"]').first.fill(password)
        log("  Campo 'Contraseña' llenado (fallback).")

    try:
        page.click('button.btn-login_v2, button[type="submit"]', timeout=5000)
        log("  Clic en 'Ingresar' enviado.")
    except Exception as e:
        raise Exception(f"No se pudo hacer clic en Ingresar: {e}")

    log("🔐 Inicio de sesión automático para Seguros Mundial. Si aparece un reCAPTCHA, resuélvelo manualmente en el navegador.")
    time.sleep(5)
    fin = time.time() + login_timeout
    while time.time() < fin:
        if job_state.get("stopping"):
            return False
        txt = _texto_pagina(page)
        if re.search(r"Inicio|DOCUMENTOS DE AYUDA|Cartera IPS|Consulta de Cartas", txt, re.I):
            log("Login exitoso, menú principal visible.")
            time.sleep(2)
            return True
        if "/home" in page.url or "/wallet" in page.url:
            log("Login exitoso por cambio de URL.")
            return True
        if "Tipo de Acceso" in txt and "Ingresar" in txt:
            log("Aún en página de login. Si hay reCAPTCHA, resuélvelo manualmente en el navegador.", "warn")
        time.sleep(2)

    log("Tiempo de espera normal agotado. Se concede 60 segundos adicionales para resolver captcha manualmente...", "warn")
    extra_time = 60
    fin_extra = time.time() + extra_time
    while time.time() < fin_extra:
        if job_state.get("stopping"):
            return False
        txt = _texto_pagina(page)
        if re.search(r"Inicio|DOCUMENTOS DE AYUDA|Cartera IPS", txt, re.I):
            log("Login exitoso después de intervención manual.")
            return True
        time.sleep(2)
    raise Exception("Tiempo de espera total agotado. No se pudo completar el login.")

# ==================== FUNCIONES PARA CONSULTA DE CARTAS ====================
def _input_por_label(page, etiqueta_regex):
    try:
        handle = page.evaluate_handle(
            """(re) => {
                const rx = new RegExp(re, 'i');
                const fields = document.querySelectorAll('mat-form-field');
                for (const f of fields) {
                    const label = f.querySelector('mat-label');
                    if (label && rx.test(label.textContent || '')) {
                        const inp = f.querySelector('input');
                        if (inp) return inp;
                    }
                }
                return null;
            }""",
            etiqueta_regex,
        )
        return handle.as_element()
    except Exception:
        return None

def _seleccionar_tipo_solicitud(page, tipo):
    inp = _input_por_label(page, r"tipo de solicitud")
    if inp is None:
        try:
            inp = page.locator('input[role="combobox"], input.mat-autocomplete-trigger').first.element_handle(timeout=4000)
        except Exception:
            inp = None
    if inp is None:
        raise Exception("No se encontró el campo 'Seleccione tipo de solicitud'.")

    inp.click()
    time.sleep(0.4)
    try:
        inp.fill("")
    except Exception:
        pass
    clave = {"No Dev/Obj": "Dev", "No Liquidación": "Liquid", "No Radicado": "Radic"}.get(tipo, tipo)
    inp.type(clave, delay=40)
    time.sleep(1.2)

    opcion_texto = tipo
    seleccionado = False
    for _ in range(3):
        try:
            clicked = page.evaluate(
                """(target) => {
                    const norm = s => (s||'').toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g,'').replace(/\\s+/g,' ').trim();
                    const t = norm(target);
                    const opts = document.querySelectorAll('mat-option, .mat-option, [role="option"]');
                    for (const o of opts) {
                        if (norm(o.textContent) === t) { o.click(); return true; }
                    }
                    for (const o of opts) {
                        if (norm(o.textContent).includes(t)) { o.click(); return true; }
                    }
                    if (opts.length === 1) { opts[0].click(); return true; }
                    return false;
                }""",
                opcion_texto,
            )
            if clicked:
                seleccionado = True
                break
        except Exception:
            pass
        time.sleep(0.8)

    if not seleccionado:
        try:
            page.locator('mat-option', has_text=re.compile(re.escape(tipo.split('/')[0].split()[-1]), re.I)).first.click(timeout=3000)
            seleccionado = True
        except Exception:
            pass

    if not seleccionado:
        raise Exception(f"No se pudo seleccionar el tipo de solicitud '{tipo}'.")
    time.sleep(0.6)
    return True

def _escribir_valor_y_consultar(page, valor):
    inp = _input_por_label(page, r"^\s*valor\s*$|valor")
    if inp is None:
        try:
            inputs = page.locator('app-input input, mat-form-field input')
            inp = inputs.nth(inputs.count() - 1).element_handle(timeout=4000)
        except Exception:
            inp = None
    if inp is None:
        raise Exception("No se encontró el campo 'Valor'.")
    inp.click()
    try:
        inp.fill("")
    except Exception:
        pass
    inp.type(valor, delay=20)
    time.sleep(0.5)

    consultado = False
    for sel in ['button:has-text("Consultar")', 'button.btn-secundario_v2']:
        try:
            page.locator(sel).first.click(timeout=4000)
            consultado = True
            break
        except Exception:
            continue
    if not consultado and not _click_por_texto(page, "Consultar", timeout=4000):
        raise Exception("No se pudo hacer clic en 'Consultar'.")
    time.sleep(2.5)

def _click_agregar(page):
    for _ in range(8):
        if job_state.get("stopping"):
            return False
        try:
            clicked = page.evaluate(
                """() => {
                    const norm = s => (s||'').trim().toLowerCase();
                    const addIcons = document.querySelectorAll('mat-icon, button, a, [role="button"]');
                    for (const el of addIcons) {
                        const txt = norm(el.textContent);
                        const title = norm(el.getAttribute('title') || '');
                        const aria = norm(el.getAttribute('aria-label') || '');
                        if (txt === 'add' || txt === 'add_circle' || txt === 'add_circle_outline' ||
                            title.includes('agregar') || aria.includes('agregar')) {
                            const target = el.closest('button, a, [role="button"]') || el;
                            target.click();
                            return true;
                        }
                    }
                    const cells = document.querySelectorAll('td, th, div[role="cell"]');
                    for (const cell of cells) {
                        if (norm(cell.textContent) === 'agregar') {
                            const btn = cell.querySelector('button, a, [role="button"]');
                            if (btn) { btn.click(); return true; }
                            else { cell.click(); return true; }
                        }
                    }
                    return false;
                }"""
            )
            if clicked:
                time.sleep(1.5)
                return True
        except Exception:
            pass
        time.sleep(1)
    raise Exception("No se pudo dar clic en AGREGAR (¿sin resultados para este valor?).")

def _abrir_buzon(page):
    for _ in range(12):
        if job_state.get("stopping"):
            return False
        try:
            estado = page.evaluate(
                """() => {
                    const icons = document.querySelectorAll('mat-icon');
                    for (const el of icons) {
                        const txt = (el.textContent||'');
                        if (txt.includes('mail_outline') || txt.includes('mail') || txt.includes('email')) {
                            const badge = el.querySelector('.mat-badge-content');
                            const n = badge ? parseInt((badge.textContent||'0').trim()||'0', 10) : 1;
                            return { found: true, count: isNaN(n) ? 0 : n };
                        }
                    }
                    return { found: false, count: 0 };
                }"""
            )
            if estado and estado.get("found") and estado.get("count", 0) >= 1:
                page.evaluate(
                    """() => {
                        const icons = document.querySelectorAll('mat-icon');
                        for (const el of icons) {
                            const txt = (el.textContent||'');
                            if (txt.includes('mail')) {
                                const target = el.closest('button, a, [role="button"]') || el;
                                target.click();
                                return true;
                            }
                        }
                        return false;
                    }"""
                )
                time.sleep(1)
                return True
        except Exception:
            pass
        time.sleep(1)
    raise Exception("El buzón no registró ninguna carta agregada (contador sigue en 0).")

def _descargar_zip_y_extraer(page, context, destino_dir, sufijo):
    """
    Descarga el ZIP del buzón, lo descomprime en destino_dir (carpeta con nombre sufijo)
    y devuelve la ruta de la carpeta donde están los PDFs.
    """
    log(f"Descargando ZIP para {sufijo}...")
    _abrir_buzon(page)

    if not _esperar_texto(page, r"Buz[oó]n de cartas|Descargar", timeout=10):
        raise Exception("No apareció el modal del buzón de cartas.")

    # Descargar el ZIP a un archivo temporal
    try:
        with page.expect_download(timeout=60000) as dl_info:
            if not _click_por_texto(page, "Descargar", timeout=8000):
                raise Exception("No se encontró el botón Descargar")
        download = dl_info.value
        # Guardar temporalmente
        temp_zip = destino_dir / f"temp_{sufijo}.zip"
        download.save_as(str(temp_zip))
        log(f"ZIP descargado: {temp_zip.name} ({temp_zip.stat().st_size // 1024} KB)")
    except Exception as e:
        log(f"Fallo descarga estándar, intentando método alternativo: {e}", "warn")
        pdf_url = None
        try:
            with context.expect_page(timeout=15000) as np_info:
                _click_por_texto(page, "Descargar", timeout=4000)
            np = np_info.value
            for _ in range(20):
                u = np.url
                if u and u != "about:blank":
                    pdf_url = u
                    break
                time.sleep(0.5)
            try:
                np.close()
            except Exception:
                pass
        except Exception:
            pass
        if pdf_url:
            resp = context.request.get(pdf_url, timeout=60000)
            if resp.ok:
                data = resp.body()
                temp_zip = destino_dir / f"temp_{sufijo}.zip"
                temp_zip.write_bytes(data)
                log(f"ZIP capturado por URL: {temp_zip.name}")
            else:
                raise Exception("No se pudo obtener el ZIP desde la URL")
        else:
            raise Exception("No se pudo descargar el ZIP del buzón")

    _click_por_texto(page, "Volver", timeout=2500)
    _click_por_texto(page, "\u00d7", timeout=1500)
    time.sleep(0.8)

    # Descomprimir en la carpeta destino
    extract_dir = destino_dir / sufijo
    extract_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(temp_zip, 'r') as zf:
        zf.extractall(extract_dir)
    log(f"ZIP descomprimido en: {extract_dir}")
    # Eliminar el ZIP temporal
    temp_zip.unlink()
    return extract_dir

def _agregar_carta_sin_tipo(page, valor, etiqueta=""):
    pref = f"[{etiqueta}] " if etiqueta else ""
    log(f"    {pref}Agregando {valor} al buzón...")
    _escribir_valor_y_consultar(page, valor)
    _click_agregar(page)
    log(f"    {pref}Carta {valor} agregada (contador +1).", "success")
    return detectar_tipo_solicitud(valor)

# ==================== AUTOMATIZACIÓN PRINCIPAL ====================
def run_automation(usuario, password, tipo_acceso, lote, valores, download_path, login_timeout):
    from playwright.sync_api import sync_playwright
    global current_browser, current_context, current_dl_dir, current_lote, current_ips_nombre

    dl_dir = Path(download_path)
    dl_dir.mkdir(parents=True, exist_ok=True)
    nit_detectado, ips_nombre = resolver_ips_por_usuario(usuario)
    if nit_detectado:
        log(f"NIT detectado en el usuario: {nit_detectado} -> IPS: {ips_nombre}")
    else:
        log("No se detecto NIT en el usuario. Carpeta: IPS_DESCONOCIDA", "warn")
    current_dl_dir = dl_dir
    current_lote = lote
    current_ips_nombre = ips_nombre
    ips_dir = dl_dir / ips_nombre
    ips_dir.mkdir(parents=True, exist_ok=True)

    # Estructura final: dentro de ips_dir tendremos DEV/ y LIQ/ (carpetas) y luego el Excel
    dev_extract_dir = ips_dir / "DEV"
    liq_extract_dir = ips_dir / "LIQ"
    dev_extract_dir.mkdir(exist_ok=True)
    liq_extract_dir.mkdir(exist_ok=True)

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--no-sandbox","--disable-dev-shm-usage"])
            context = browser.new_context(accept_downloads=True, viewport={"width": 1500, "height": 900})
            page = context.new_page()
            current_browser = browser
            current_context = context

            if not _hacer_login(page, usuario, password, login_timeout):
                if job_state.get("stopping"):
                    return
            if job_state.get("stopping"):
                return

            log("Navegando directamente a Consulta de Cartas...")
            page.goto(CONSULTA_CARTAS_URL, wait_until="domcontentloaded", timeout=60000)
            time.sleep(3)
            if not _esperar_texto(page, r"Seleccione tipo de solicitud", timeout=15):
                raise Exception("No se detectó la pantalla de Consulta de Cartas después de navegación directa.")
            log("Pantalla 'Consulta de Cartas' cargada correctamente.")

            completadas = cargar_progreso(ips_dir)
            pendientes = []
            for v in valores:
                if v in completadas:
                    log(f"Omitiendo (ya procesada): {v}")
                    with job_lock:
                        job_state["stats"]["descargadas"] += 1
                else:
                    pendientes.append(v)
            with job_lock:
                job_state["stats"]["total"] = len(valores)

            lista_dev = [v for v in pendientes if detectar_tipo_solicitud(v) == "No Dev/Obj"]
            lista_liq = [v for v in pendientes if detectar_tipo_solicitud(v) == "No Liquidación"]
            otros = [v for v in pendientes if v not in lista_dev and v not in lista_liq]
            for v in otros:
                lista_dev.append(v)

            log(f"Total a procesar: {len(pendientes)} | DEV: {len(lista_dev)} | LIQ: {len(lista_liq)}")

            # Procesar DEV (agregar cartas y luego descargar ZIP y extraer)
            if lista_dev:
                log(f"=== Procesando {len(lista_dev)} cartas DEV/Obj ===")
                log("Seleccionando tipo 'No Dev/Obj' (solo una vez)...")
                _seleccionar_tipo_solicitud(page, "No Dev/Obj")
                # Registrar cartas para el Excel posterior
                cartas_dev_registradas = []
                for idx, valor in enumerate(lista_dev, 1):
                    if job_state.get("stopping"):
                        break
                    log(f"[DEV {idx}/{len(lista_dev)}]")
                    try:
                        _agregar_carta_sin_tipo(page, valor, etiqueta="DEV")
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        with job_lock:
                            job_state["descargas_exitosas"].append({
                                "consecutivo": valor,
                                "tipo": "No Dev/Obj",
                                "archivo": "",  # se llenará después con la ruta dentro del ZIP final
                                "timestamp": timestamp,
                            })
                            job_state["stats"]["descargadas"] += 1
                        completadas.add(valor)
                        guardar_progreso(ips_dir, completadas)
                        cartas_dev_registradas.append({
                            "consecutivo": valor,
                            "tipo": "No Dev/Obj",
                            "timestamp": timestamp
                        })
                    except Exception as e:
                        log(f"  Error al agregar {valor}: {e}", "error")
                        with job_lock:
                            job_state["errores_detalle"].append({
                                "consecutivo": valor,
                                "tipo": "No Dev/Obj",
                                "error": str(e),
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "captura": "",
                            })
                            job_state["stats"]["errores"] += 1
                        try:
                            errores_dir = ips_dir / "Errores"
                            errores_dir.mkdir(parents=True, exist_ok=True)
                            cap = errores_dir / f"ERROR_{re.sub(r'[^A-Za-z0-9_-]','_', valor)}.png"
                            page.screenshot(path=str(cap))
                            with job_lock:
                                job_state["errores_detalle"][-1]["captura"] = str(cap)
                        except Exception:
                            pass

                if not job_state.get("stopping") and lista_dev:
                    log("Descargando y extrayendo ZIP de DEV...")
                    try:
                        _descargar_zip_y_extraer(page, context, ips_dir, "DEV")
                        # Las cartas ahora están en ips_dir/DEV/
                        # Actualizar las rutas en job_state para el Excel final
                        with job_lock:
                            for ex in job_state["descargas_exitosas"]:
                                if ex["tipo"] == "No Dev/Obj" and not ex["archivo"]:
                                    # Asignar una ruta simbólica dentro del ZIP final
                                    ex["archivo"] = f"DEV/{ex['consecutivo']}.pdf"  # aproximado, solo para referencia
                    except Exception as e:
                        log(f"Error al procesar ZIP de DEV: {e}", "error")

            # Procesar LIQ
            if lista_liq and not job_state.get("stopping"):
                log(f"=== Procesando {len(lista_liq)} cartas LIQ ===")
                # Recargar la página para resetear el buzón
                page.goto(CONSULTA_CARTAS_URL, wait_until="domcontentloaded", timeout=60000)
                time.sleep(2)
                if not _esperar_texto(page, r"Seleccione tipo de solicitud", timeout=10):
                    raise Exception("No se detectó la pantalla de Consulta de Cartas para LIQ.")
                log("Seleccionando tipo 'No Liquidación' (solo una vez)...")
                _seleccionar_tipo_solicitud(page, "No Liquidación")
                cartas_liq_registradas = []
                for idx, valor in enumerate(lista_liq, 1):
                    if job_state.get("stopping"):
                        break
                    log(f"[LIQ {idx}/{len(lista_liq)}]")
                    try:
                        _agregar_carta_sin_tipo(page, valor, etiqueta="LIQ")
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        with job_lock:
                            job_state["descargas_exitosas"].append({
                                "consecutivo": valor,
                                "tipo": "No Liquidación",
                                "archivo": "",
                                "timestamp": timestamp,
                            })
                            job_state["stats"]["descargadas"] += 1
                        completadas.add(valor)
                        guardar_progreso(ips_dir, completadas)
                        cartas_liq_registradas.append({
                            "consecutivo": valor,
                            "tipo": "No Liquidación",
                            "timestamp": timestamp
                        })
                    except Exception as e:
                        log(f"  Error al agregar {valor}: {e}", "error")
                        with job_lock:
                            job_state["errores_detalle"].append({
                                "consecutivo": valor,
                                "tipo": "No Liquidación",
                                "error": str(e),
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "captura": "",
                            })
                            job_state["stats"]["errores"] += 1
                        try:
                            errores_dir = ips_dir / "Errores"
                            errores_dir.mkdir(parents=True, exist_ok=True)
                            cap = errores_dir / f"ERROR_{re.sub(r'[^A-Za-z0-9_-]','_', valor)}.png"
                            page.screenshot(path=str(cap))
                            with job_lock:
                                job_state["errores_detalle"][-1]["captura"] = str(cap)
                        except Exception:
                            pass

                if not job_state.get("stopping") and lista_liq:
                    log("Descargando y extrayendo ZIP de LIQ...")
                    try:
                        _descargar_zip_y_extraer(page, context, ips_dir, "LIQ")
                        with job_lock:
                            for ex in job_state["descargas_exitosas"]:
                                if ex["tipo"] == "No Liquidación" and not ex["archivo"]:
                                    ex["archivo"] = f"LIQ/{ex['consecutivo']}.pdf"
                    except Exception as e:
                        log(f"Error al procesar ZIP de LIQ: {e}", "error")

            if job_state.get("stopping"):
                browser.close()
                return

            browser.close()

            # Ahora crear el Excel general con todas las cartas
            with job_lock:
                exitosas = job_state["descargas_exitosas"].copy()
                errores = job_state["errores_detalle"].copy()
            if EXCEL_AVAILABLE:
                excel_path = ips_dir / "reporte_cartas.xlsx"
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Cartas procesadas"
                ws.append(["Consecutivo", "Tipo solicitud", "Carpeta", "Fecha/Hora"])
                for ex in exitosas:
                    carpeta = "DEV" if ex["tipo"] == "No Dev/Obj" else "LIQ"
                    ws.append([ex.get("consecutivo"), ex.get("tipo"), carpeta, ex.get("timestamp")])
                if errores:
                    wse = wb.create_sheet("Errores")
                    wse.append(["Consecutivo", "Tipo solicitud", "Error", "Captura", "Fecha/Hora"])
                    for er in errores:
                        wse.append([er.get("consecutivo"), er.get("tipo"), er.get("error"), er.get("captura"), er.get("timestamp")])
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                wb.save(excel_path)
                log(f"Reporte Excel generado: {excel_path.name}")
            else:
                excel_path = None

            # Crear el ZIP único final con nombre de la IPS
            final_zip_path = dl_dir / f"{ips_nombre}.zip"
            with zipfile.ZipFile(final_zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                # Agregar carpeta DEV
                if dev_extract_dir.exists():
                    for file_path in dev_extract_dir.rglob("*"):
                        if file_path.is_file():
                            arcname = f"DEV/{file_path.relative_to(dev_extract_dir)}"
                            zf.write(file_path, arcname)
                # Agregar carpeta LIQ
                if liq_extract_dir.exists():
                    for file_path in liq_extract_dir.rglob("*"):
                        if file_path.is_file():
                            arcname = f"LIQ/{file_path.relative_to(liq_extract_dir)}"
                            zf.write(file_path, arcname)
                # Agregar Excel
                if excel_path and excel_path.exists():
                    zf.write(excel_path, "reporte_cartas.xlsx")
                # Agregar carpeta de errores si existe
                errores_dir = ips_dir / "Errores"
                if errores_dir.exists():
                    for file_path in errores_dir.rglob("*"):
                        if file_path.is_file():
                            arcname = f"Errores/{file_path.relative_to(errores_dir)}"
                            zf.write(file_path, arcname)
            log(f"ZIP final creado: {final_zip_path.name}", "success")
            # Limpiar las carpetas extraídas (opcional, para no duplicar espacio)
            shutil.rmtree(dev_extract_dir, ignore_errors=True)
            shutil.rmtree(liq_extract_dir, ignore_errors=True)
            if excel_path:
                excel_path.unlink()
            # El progreso.json lo dejamos

            total_errores = len(errores)
            if total_errores:
                log(f"Proceso completado con {total_errores} error(es). Ver Excel en el ZIP.", "warn")
            else:
                log("Proceso completado sin errores.", "success")

    except Exception as e:
        if not job_state.get("stopping"):
            log(f"Error critico: {e}", "error")
            with job_lock:
                job_state["error"] = str(e)
        else:
            log("Proceso detenido por el usuario.")
    finally:
        with job_lock:
            job_state["running"] = False
            job_state["finished"] = True
            job_state["stopping"] = False
        current_browser = None
        current_context = None
        current_dl_dir = None
        current_lote = None
        current_ips_nombre = None

# ==================== PARSEO DE VALORES ====================
def parse_valores(texto):
    if not texto:
        return []
    crudos = re.split(r"[\s,;]+", texto.strip())
    out = [c.strip() for c in crudos if c.strip()]
    vistos = set()
    res = []
    for c in out:
        if c not in vistos:
            vistos.add(c)
            res.append(c)
    return res

# ==================== RUTAS FLASK ====================
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/start", methods=["POST"])
def start_job():
    data = request.json or {}
    usuario = data.get("usuario", "").strip()
    password = data.get("password", "").strip()
    tipo_acceso = data.get("tipo_acceso", "").strip()
    lote = data.get("lote", "").strip() or datetime.now().strftime("Lote_%Y%m%d_%H%M")
    valores_texto = data.get("valores", "").strip()
    custom_path = data.get("download_path", "").strip()
    login_timeout = int(data.get("login_timeout", 180) or 180)

    if not all([usuario, password]):
        return jsonify({"ok": False, "error": "Faltan usuario o contraseña"}), 400

    valores = parse_valores(valores_texto)
    if not valores:
        return jsonify({"ok": False, "error": "No hay consecutivos para procesar. Carga el Excel o pega la lista."}), 400

    lote_safe = re.sub(r"[^\w\-]", "_", lote)

    with job_lock:
        if job_state["running"]:
            return jsonify({"ok": False, "error": "Ya hay un proceso en ejecucion"}), 409
        job_state["running"] = True
        job_state["finished"] = False
        job_state["error"] = None
        job_state["stats"] = {"total": len(valores), "descargadas": 0, "errores": 0}
        job_state["errores_detalle"] = []
        job_state["descargas_exitosas"] = []
        job_state["logs"] = []

    dl_path = custom_path if custom_path else str(DOWNLOAD_DIR / lote_safe)
    t = threading.Thread(
        target=run_automation,
        args=(usuario, password, tipo_acceso, lote_safe, valores, dl_path, login_timeout),
        daemon=True,
    )
    t.start()
    return jsonify({"ok": True, "download_path": dl_path, "total": len(valores), "lote": lote_safe})

@app.route("/api/stop", methods=["POST"])
def stop_job_route():
    with job_lock:
        if not job_state["running"]:
            return jsonify({"ok": False, "message": "No hay proceso en ejecucion"}), 400
    stop_job()
    return jsonify({"ok": True, "message": "Deteniendo proceso..."})

@app.route("/api/reset", methods=["POST"])
def reset_job_route():
    data = request.json or {}
    lote = data.get("lote", "").strip()
    with job_lock:
        if job_state["running"]:
            stop_job()
            time.sleep(2)
    if lote:
        lote_safe = re.sub(r"[^\w\-]", "_", lote)
        lote_dir = DOWNLOAD_DIR / lote_safe
        borrados = 0
        if lote_dir.exists():
            for cand in list(lote_dir.glob("**/progreso.json")):
                try:
                    cand.unlink()
                    borrados += 1
                    log(f"Progreso eliminado: {cand}")
                except Exception as e:
                    log(f"Error al borrar progreso: {e}", "warn")
        if borrados == 0:
            log("No se encontro progreso para borrar en ese lote.", "warn")
    reset_state()
    return jsonify({"ok": True, "message": "Estado reiniciado y progreso eliminado."})

@app.route("/api/status")
def get_status():
    with job_lock:
        return jsonify({
            "running": job_state["running"],
            "finished": job_state["finished"],
            "error": job_state["error"],
            "stats": job_state["stats"],
        })

@app.route("/api/logs")
def get_logs():
    since = int(request.args.get("since", 0))
    with job_lock:
        return jsonify({"logs": job_state["logs"][since:]})

@app.route("/api/logs", methods=["DELETE"])
def clear_logs():
    with job_lock:
        job_state["logs"] = []
    return jsonify({"ok": True})

@app.route("/api/files")
def list_files():
    lote = request.args.get("lote", "")
    folder = DOWNLOAD_DIR / lote if lote else DOWNLOAD_DIR
    files = []
    if folder.exists():
        # Buscar el ZIP final con nombre de IPS (no los ZIPs temporales)
        for zip_file in folder.rglob("*.zip"):
            # Excluir cualquier ZIP que no sea el final (por ejemplo, los que contengan "temp" o "PARCIAL")
            if "temp_" in zip_file.name or "PARCIAL" in zip_file.name:
                continue
            # También evitar el ZIP completo de lote si se generara (pero ya no)
            files.append({
                "name": zip_file.name,
                "size": zip_file.stat().st_size,
                "lote": lote,
                "path": zip_file.name
            })
    return jsonify({"files": files})

@app.route("/api/files", methods=["DELETE"])
def delete_all_files():
    lote = request.args.get("lote", "")
    folder = DOWNLOAD_DIR / lote if lote else DOWNLOAD_DIR
    if not folder.exists():
        return jsonify({"ok": True, "message": "No hay archivos que eliminar"})
    try:
        eliminados = 0
        for item in list(folder.iterdir()):
            if item.is_file():
                item.unlink()
                eliminados += 1
            elif item.is_dir():
                for sub in item.rglob("*"):
                    if sub.is_file() and sub.name != "progreso.json":
                        sub.unlink()
                        eliminados += 1
        log(f"Soportes eliminados: {eliminados} (progreso conservado)")
        return jsonify({"ok": True, "message": f"Se eliminaron {eliminados} archivo(s). El progreso se conservo.", "eliminados": eliminados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/files/soportes", methods=["DELETE"])
def delete_soportes():
    """Borra ZIPs, PDFs y carpetas de soportes pero conserva progreso.json intacto."""
    lote = request.args.get("lote", "")
    folder = DOWNLOAD_DIR / lote if lote else DOWNLOAD_DIR
    if not folder.exists():
        return jsonify({"ok": True, "message": "No hay soportes que eliminar", "eliminados": 0})
    CONSERVAR = {"progreso.json"}
    try:
        eliminados = 0
        for path in sorted(folder.rglob("*"), reverse=True):
            if path.name in CONSERVAR:
                continue
            if path.is_file():
                path.unlink()
                eliminados += 1
            elif path.is_dir():
                try:
                    path.rmdir()  # solo borra si quedó vacía
                except OSError:
                    pass
        log(f"Soportes eliminados: {eliminados} archivos. progreso.json conservado.")
        return jsonify({"ok": True, "message": f"{eliminados} soporte(s) eliminado(s). El progreso se conservó.", "eliminados": eliminados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/progreso")
def get_progreso():
    """Devuelve el historial de progreso de todos los lotes guardados en disco."""
    lote = request.args.get("lote", "")
    resultados = []
    base = DOWNLOAD_DIR / lote if lote else DOWNLOAD_DIR
    for prog_file in sorted(base.rglob("progreso.json")):
        try:
            data = json.loads(prog_file.read_text(encoding="utf-8"))
            # Reconstruir ruta relativa para mostrar: lote/IPS
            partes = prog_file.parent.relative_to(DOWNLOAD_DIR).parts
            resultados.append({
                "ruta": "/".join(partes),
                "lote": partes[0] if partes else "",
                "ips": partes[1] if len(partes) > 1 else "",
                "completadas": sorted(data) if isinstance(data, list) else sorted(data.get("completadas", data) if isinstance(data, dict) else []),
                "total": len(data) if isinstance(data, (list, set)) else len(data.get("completadas", data) if isinstance(data, dict) else []),
            })
        except Exception:
            pass
    return jsonify({"progreso": resultados})

@app.route("/downloads/<path:filename>")
def download_file(filename):
    for file_path in DOWNLOAD_DIR.rglob(filename):
        if file_path.is_file():
            return send_from_directory(file_path.parent, file_path.name, as_attachment=True)
    return jsonify({"error": "Archivo no encontrado"}), 404

@app.route("/api/upload", methods=["POST"])
def upload_consecutivos():
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "No se envio ningun archivo"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"ok": False, "error": "Archivo vacio"}), 400
    try:
        filename = file.filename.lower()
        valores = []
        if filename.endswith('.csv'):
            raw = file.read()
            for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
                try:
                    txt = raw.decode(enc)
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            else:
                txt = raw.decode('latin-1', errors='replace')
            reader = csv.reader(txt.splitlines())
            rows = list(reader)
            header = rows[0] if rows else []
            col_idx = 0
            for i, h in enumerate(header):
                hl = (h or '').lower()
                if 'consecutiv' in hl or 'radicad' in hl:
                    col_idx = i
                    break
            start = 1 if header and not re.search(r'^(DEV|LIQ|CMV)', (header[col_idx] or '').upper()) else 0
            for r in rows[start:]:
                if len(r) > col_idx and r[col_idx].strip():
                    valores.append(r[col_idx].strip())
        elif filename.endswith(('.xls', '.xlsx')):
            if not EXCEL_AVAILABLE:
                return jsonify({"ok": False, "error": "openpyxl no instalado"}), 500
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            ws = wb.active
            col_idx = 1
            found = False
            for cell in ws[1]:
                if cell.value and re.search(r'consecutiv|radicad', str(cell.value), re.I):
                    col_idx = cell.column
                    found = True
                    break
            start_row = 2 if found else 1
            first = ws.cell(row=1, column=col_idx).value
            if first and re.match(r'^(DEV|LIQ|CMV)', str(first).upper()):
                start_row = 1
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                val = row[col_idx - 1] if len(row) >= col_idx else None
                if val and str(val).strip():
                    valores.append(str(val).strip())
        else:
            return jsonify({"ok": False, "error": "Formato no soportado. Use CSV o Excel"}), 400

        vistos = set()
        limpios = []
        for v in valores:
            v = v.strip()
            if v and v not in vistos:
                vistos.add(v)
                limpios.append(v)
        if not limpios:
            return jsonify({"ok": False, "error": "No se encontraron consecutivos validos en el archivo"}), 400
        resumen = {}
        for v in limpios:
            t = detectar_tipo_solicitud(v)
            resumen[t] = resumen.get(t, 0) + 1
        return jsonify({"ok": True, "count": len(limpios), "valores": limpios, "resumen": resumen})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al procesar archivo: {str(e)}"}), 500

if __name__ == "__main__":
    print("\n" + "=" * 55)
    print("  Descargador de Cartas Glosa - Seguros Mundial")
    print("  Portal A3M / iqdigital (un solo ZIP por IPS con carpetas DEV/LIQ y Excel)")
    print("=" * 55)
    print(f"  Carpeta de descargas: {DOWNLOAD_DIR}")
    print(f"  Puerto: {port}")
    print("=" * 55 + "\n")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)